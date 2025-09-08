import os
import threading
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

#from sentence_transformers import SentenceTransformer, util
#model = SentenceTransformer('all-MiniLM-L6-v2')

folder = "C:\\Users\\AdminAkro\\Downloads\\"
file_path1 = "C:\\Users\\AdminAkro\\Downloads\\[WMATA] Trigger Condition Faults List\\FE010019308_FaultsList.xlsx"
col_name1 ='TriggerCondition'
file_path2 = "C:\\Users\\AdminAkro\\Downloads\\[WMATA] Trigger Condition Faults List\\FE010019308_FaultsList_old.xlsx"
col_name2 ='TriggerCondition'
file_path_output = "C:\\Users\\AdminAkro\\Downloads\\out.xlsx"
file_path_input_to_clone = "C:\\Users\\AdminAkro\\Downloads\\FE010019308_FaultsList.xlsx"
file_path_output_updated = "C:\\Users\\AdminAkro\\Downloads\\FE010019308_FaultsList_updated.xlsx"
col3 = []
threshold = 0.3
index_row_to_cut = 9

def get_embedding_cached(text):
    print(f"Getting embedding for text: {text}")
    #Get embedding with thread-safe cache
    with cache_lock:
        if text not in embeddings_cache:
            embeddings_cache[text] = model.encode(text, convert_to_tensor=True)
        return embeddings_cache[text]

def are_conceptually_similar(sentence1, sentence2):
    if sentence1 == sentence2:
        return True
    else:
        return False
    # You can uncomment this code to enable IA similarity check
    #embedding1 = get_embedding_cached(sentence1)
    #embedding2 = get_embedding_cached(sentence2)
    #cosine_similarity = util.pytorch_cos_sim(embedding1, embedding2).item()
    #return cosine_similarity

def read_excel_column(xls, col_name):
    try:
        column = xls[col_name]
        col = column.astype(str).tolist()
        return col
    except Exception as e:
        print(f"column index out of range. I'm going on with the next column. Error: {e}")
        col = ["", ""]
        return col

def read_excel_row(xls, row_index):
    try:
        riga = xls.iloc[row_index -1]
        r = riga.astype(str).tolist()
        return r
    except Exception as e:
        print(f"row index out of range. I'm going on with the next row. Error: {e}")
        r = ["", ""]
        return r

def list_of_sheets(file_path):
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    sheet_names = xls.sheet_names
    return sheet_names

def write_excel_with_column(file_path, column_to_write, sheet_name):
    df = pd.DataFrame({col_name1 :column_to_write})
    write_excel(df, file_path, sheet_name)

def write_excel(df, file_path, sheet_name):
    try:
        if not os.path.exists(file_path):
            mode = Literal('w')
        else:
            mode = Literal('a')

        with pd.ExcelWriter(file_path, mode=mode, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            # Set wrap text on all cells in the sheet
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.iter_cols(min_col=1, max_col=worksheet.max_column):
                col_index = col[0].column
                for cell in col:
                    cell.alignment = Alignment(wrapText=True)
                worksheet.column_dimensions[get_column_letter(col_index)].width = 50
        print(f"I wrote the file {file_path} in {sheet_name}")
    except Exception as e:
        print(f"Error writing the file {file_path}: {e}")

def cell_union(cell1, cell2):
    lines1 = str(cell1).split('\n')
    lines2 = str(cell2).split('\n')
    if not lines1[0].strip() == lines2[0].strip() and not are_conceptually_similar(cell1, cell2):
        string = f"{cell1} \n {cell2}"
    else:
        string = f"{cell1}"
    return string

def number_of_row(xls):
    try:
        return xls.shape[0]
    except Exception as e:
        print(f"Error counting row: {e}")
        return -1

def get_cell(xls, row_index, col_index):
    try:
        cell_value = xls.at[row_index, col_index]
        return str(cell_value)
    except Exception as e:
        print(f"Error reading cell at row {row_index}, column {col_index}: {e}")
        return None

def cut_row(row, index_row_cut):
    return row[:index_row_cut]

def there_is_a_column(xls, sheet_name, col_name):
    try:
        col_index = xls.columns.tolist().index(col_name)
        return True
    except Exception as e:
        print(f"The page {sheet_name} doesn't have a column named {col_name}: {e}")
        return False

def create_complete_file(column_to_write):
    sheet_names = list_of_sheets(file_path_output_new)
    for sheet_name in sheet_names:
        xls = pd.read_excel(file_path_output_new, sheet_name=sheet_name, header=0)
        if sheet_name == sheet_name1 and there_is_a_column(xls, sheet_name, col_name1):
            xls[col_name1] = column_to_write
            print(xls[col_name1])
            write_excel(xls, file_path_output_updated, sheet_name)

def process_sheets(xls1, xls2):
    bucket_of_analyzed_rows = []
    for row_index1 in range(0, row_number1):
        row1 = cut_row(read_excel_row(xls1, row_index1), index_row_to_cut)
        for row_index2 in range(0, row_number2):
            row2 = cut_row(read_excel_row(xls2, row_index2), index_row_to_cut)
            cell1 = get_cell(xls1, row_index1, col_name1)
            if row1 == row2:
                cell2 = get_cell(xls2, row_index2, col_name2)
                cell3 = cell_union(cell1, cell2)
                bucket_of_analyzed_rows.append(row1)
                col3.append(cell3)
        if row1 not in bucket_of_analyzed_rows:
            cell1 = get_cell(xls1, row_index1, col_name1)
            cell3 = cell1
            col3.append(cell3)
        else:
            continue

if __name__ == '__main__':
        sheet_names1 = list_of_sheets(file_path1)
        sheet_names2 = list_of_sheets(file_path2)

        if os.path.exists(file_path_output):
            os.remove(file_path_output)

        for sheet_name1 in sheet_names1:
            for sheet_name2 in sheet_names2:
                xls1 = pd.read_excel(file_path1, sheet_name=sheet_name1, header=0)
                xls2 = pd.read_excel(file_path2, sheet_name=sheet_name2, header=0)
                row_number1 = number_of_row(xls1)
                row_number2 = number_of_row(xls2)
                # If sheets have the same name and there is a column named col_name1
                if sheet_name1 == sheet_name2 and there_is_a_column(xls1, sheet_name1, col_name1) and there_is_a_column(xls2, sheet_name2, col_name2):
                    print(f"Analyzing {sheet_name1} and {sheet_name2}")
                    # For on row, grab the first index_row_to_cut cells, get the cell
                    # at position row_index with col_name.
                    # If the cuted rows are the same, copy the cell1's value and cell2's value
                    # in a new cell3. This cell3s will populate col3
                    process_sheets(xls1, xls2)
                    # write col3 on a file in a sheet named sheet_name1. If the file doesn't exist, create it
                    write_excel_with_column(file_path_output, col3, sheet_name1)
                    # create another new file contains all the same data as file_path_1 except
                    # the column named col_name1 which is substituted with col3.
                    create_complete_file(col3)
                    # delete col3 and go to the next sheet
                    col3 = []
                else:
                    continue
        print("I've finished")